import json
import re
from pathlib import Path

import openpyxl


_MAPPINGS_SHEET = "mappings"
_CONFIG_SHEETS = {"config", "_config"}
_COMPUTED_PREFIX = "(computed)"


class MappingLoader:
    def __init__(self, mapping_path):
        self.mapping_path = mapping_path
        self._extension = Path(mapping_path).suffix.lower()

    def load(self) -> dict:
        """Return only the field mapping rules (excludes the 'config' key)."""
        raw = self._load_raw()
        rules = {k: v for k, v in raw.items() if k != "config"}
        config = raw.get("config", {})
        if "file_name" in rules and isinstance(rules["file_name"], dict):
            if isinstance(config, dict) and "name" in config:
                rules["file_name"] = {**rules["file_name"], "name": config["name"]}
            else:
                derived_name = self._derive_file_name_name(rules["file_name"])
                if derived_name:
                    rules["file_name"] = {**rules["file_name"], "name": derived_name}
        return rules

    def load_config(self) -> dict:
        """Return the 'config' block, or an empty dict if absent."""
        return self._load_raw().get("config", {})

    def update_config(self, **kwargs) -> None:
        """
        Update specific keys inside the 'config' block and write back to disk.
        Example: update_config(data_file="C:/Reports/table.xlsx")
        """
        raw = self._load_raw()
        config = raw.setdefault("config", {})
        if not isinstance(config, dict):
            config = {}
            raw["config"] = config
        config.update(kwargs)
        self._save_raw(raw)

    def update_file_name_field(self, **kwargs) -> None:
        """
        Update specific keys inside the 'file_name' rule and write back to disk.
        Example: update_file_name_field(name="NOVOCIB Rapport d'essai")
        """
        raw = self._load_raw()

        if self._extension == ".json":
            if "file_name" not in raw or not isinstance(raw["file_name"], dict):
                raw["file_name"] = {"operation": "format", "format": "{name} {numero_rapport}.docx"}
            raw["file_name"].update(kwargs)
        else:
            config = raw.setdefault("config", {})
            if not isinstance(config, dict):
                config = {}
                raw["config"] = config
            config.update(kwargs)

        self._save_raw(raw)

    def load_file_name_field(self) -> dict:
        """Return the 'file_name' rule dict, or empty dict if absent."""
        raw = self._load_raw()
        file_name = raw.get("file_name", {})
        if not isinstance(file_name, dict):
            return {}
        if isinstance(raw.get("config"), dict) and "name" in raw["config"]:
            return {**file_name, "name": raw["config"]["name"]}
        derived_name = self._derive_file_name_name(file_name)
        if derived_name:
            return {**file_name, "name": derived_name}
        return file_name

    def _derive_file_name_name(self, file_name_rule: dict) -> str | None:
        if not isinstance(file_name_rule, dict):
            return None
        if file_name_rule.get("operation") != "format":
            return None
        fmt = file_name_rule.get("format")
        if not isinstance(fmt, str):
            return None
        if "{name}" in fmt:
            return None
        name = fmt
        if name.lower().endswith(".docx"):
            name = name[:-5]
        return name.strip()

    def _load_raw(self) -> dict:
        if self._extension == ".json":
            return self._load_json()
        if self._extension == ".xlsx":
            return self._load_xlsx()
        raise ValueError(f"Unsupported mapping file type: {self._extension}")

    def _save_raw(self, raw: dict) -> None:
        if self._extension == ".json":
            with open(self.mapping_path, "w", encoding="utf-8") as f:
                json.dump(raw, f, indent=2, ensure_ascii=False)
            return

        if self._extension == ".xlsx":
            self._save_xlsx(raw)
            return

        raise ValueError(f"Unsupported mapping file type: {self._extension}")

    def _load_json(self) -> dict:
        with open(self.mapping_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _find_sheet(self, workbook, sheet_name: str):
        for name in workbook.sheetnames:
            if name.lower() == sheet_name.lower():
                return workbook[name]
        return None

    def _load_xlsx(self) -> dict:
        workbook = openpyxl.load_workbook(self.mapping_path, data_only=True)
        sheet = self._find_sheet(workbook, _MAPPINGS_SHEET)
        if sheet is None:
            raise ValueError(f"Workbook must contain a '{_MAPPINGS_SHEET}' sheet.")

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {}

        headers = [self._normalize_header(value).lower() for value in rows[0]]
        header_index = {header: index for index, header in enumerate(headers) if header}

        result = {}
        for row in rows[1:]:
            if not row:
                continue
            column_value = self._safe_cell(row, header_index.get("spreadsheet column"))
            if not column_value:
                continue

            column_text = str(column_value).strip()
            is_computed = column_text.lower().startswith(_COMPUTED_PREFIX)
            key = column_text[len(_COMPUTED_PREFIX):].strip() if is_computed else column_text
            if not key:
                continue

            placeholder_value = self._safe_cell(row, header_index.get("placeholder"))
            placeholder = self._normalize_header(placeholder_value)
            if placeholder in {"", "—", "-"}:
                placeholder = None

            operation_value = self._safe_cell(row, header_index.get("operation"))
            operation_text = str(operation_value).strip() if operation_value is not None else ""

            if is_computed:
                rule = self._parse_computed_rule(operation_text)
            else:
                rule = {"column": key}
                if placeholder:
                    rule["placeholder"] = placeholder
                operations = self._parse_operations(operation_text)
                if operations:
                    rule["operations"] = operations

            if not rule:
                continue
            result[key] = rule

        config = self._load_xlsx_config(workbook)
        if config:
            if "file_name" in result and isinstance(result["file_name"], dict) and "name" in config:
                result["file_name"] = {**result["file_name"], "name": config["name"]}
            result["config"] = {k: v for k, v in config.items() if k != "name"}
        return result

    def _safe_cell(self, row, index):
        if index is None or index >= len(row):
            return None
        return row[index]

    def _load_xlsx_config(self, workbook) -> dict:
        config_sheet = None
        for candidate in _CONFIG_SHEETS:
            config_sheet = self._find_sheet(workbook, candidate)
            if config_sheet is not None:
                break
        if config_sheet is None:
            return {}

        rows = list(config_sheet.iter_rows(values_only=True))
        if not rows:
            return {}

        start_index = 0
        first_row = rows[0]
        first_key = self._normalize_header(first_row[0]).lower() if first_row and first_row[0] is not None else ""
        first_value = self._normalize_header(first_row[1]).lower() if first_row and len(first_row) > 1 and first_row[1] is not None else ""
        if first_key in {"key", "name"} and first_value in {"value", "val"}:
            start_index = 1

        config = {}
        for row in rows[start_index:]:
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            config[key] = value
        return config

    def _save_xlsx(self, raw: dict) -> None:
        workbook = openpyxl.load_workbook(self.mapping_path)
        config_sheet = None
        for candidate in _CONFIG_SHEETS:
            if candidate in workbook.sheetnames:
                config_sheet = workbook[candidate]
                break
        if config_sheet is None:
            config_sheet = workbook.create_sheet("_config")
            config_sheet.sheet_state = "hidden"

        config_sheet.delete_rows(1, config_sheet.max_row)
        config = raw.get("config", {})
        if isinstance(config, dict):
            for row_index, (key, value) in enumerate(config.items(), start=1):
                config_sheet.cell(row=row_index, column=1, value=key)
                config_sheet.cell(row=row_index, column=2, value=value)

        workbook.save(self.mapping_path)

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _parse_operations(self, text: str) -> list[dict] | None:
        if not text:
            return None
        parts = [part.strip() for part in text.split(";") if part.strip()]
        ops = []
        for part in parts:
            lower = part.lower()
            if lower == "formula":
                ops.append({"type": "formula"})
                continue
            if lower.startswith("multiply "):
                ops.append({"type": "multiply", "value": self._parse_value(part[9:].strip())})
                continue
            if lower.startswith("divide "):
                ops.append({"type": "divide", "value": self._parse_value(part[7:].strip())})
                continue
            if lower.startswith("add "):
                ops.append({"type": "add", "value": self._parse_value(part[4:].strip())})
                continue
            if lower.startswith("subtract "):
                ops.append({"type": "subtract", "value": self._parse_value(part[9:].strip())})
                continue
            if lower.startswith("round(") and part.endswith(")"):
                number = part[6:-1].strip()
                ops.append({"type": "round", "decimals": int(number) if number.isdigit() else 0})
                continue
            if lower.startswith("round "):
                ops.append({"type": "round", "decimals": int(part[6:].strip()) if part[6:].strip().isdigit() else 0})
                continue
            if lower.startswith("suffix "):
                ops.append({"type": "suffix", "value": self._parse_value(part[7:].strip())})
                continue
            if lower.startswith("prefix "):
                ops.append({"type": "prefix", "value": self._parse_value(part[7:].strip())})
                continue
            if lower.startswith("date_format "):
                ops.append({"type": "date_format", "format": part[12:].strip()})
                continue
            if lower in {"upper", "uppercase"}:
                ops.append({"type": "upper"})
                continue
            if lower in {"lower", "lowercase"}:
                ops.append({"type": "lower"})
                continue
            if lower == "strip":
                ops.append({"type": "strip"})
                continue
            if lower.startswith("upper(") and part.endswith(")"):
                ops.append({"type": "upper"})
                continue
            if lower.startswith("lower(") and part.endswith(")"):
                ops.append({"type": "lower"})
                continue
            if lower.startswith("strip(") and part.endswith(")"):
                ops.append({"type": "strip"})
                continue
        return ops or None

    def _parse_computed_rule(self, text: str) -> dict:
        text = text.strip()
        if not text:
            return {}

        op_name, payload = self._split_operation_text(text)
        op_name = op_name.lower()
        op_name = self._normalize_computed_operation_name(op_name)

        rule: dict = {"operation": op_name}
        if op_name == "format":
            if payload:
                rule["format"] = self._parse_value(payload)
            return rule

        if op_name == "today":
            if payload:
                if payload.lower().startswith("format "):
                    rule["format"] = payload[7:].strip()
                else:
                    rule["format"] = payload
            return rule

        if op_name == "excel_day_counter":
            self._merge_keyed_params(rule, payload)
            return rule

        if op_name == "date_format":
            rule["format"] = payload
            return rule

        self._merge_keyed_params(rule, payload)
        return rule

    def _split_operation_text(self, text: str) -> tuple[str, str]:
        if "(" in text and text.endswith(")"):
            op_name, rest = text.split("(", 1)
            return op_name.strip(), rest[:-1].strip()

        parts = text.split(None, 1)
        if len(parts) == 1:
            return parts[0], ""
        return parts[0], parts[1]

    @staticmethod
    def _normalize_computed_operation_name(name: str) -> str:
        if name in {"report_number", "repport_number", "numero_rapport"}:
            return "excel_day_counter"
        return name

    def _merge_keyed_params(self, rule: dict, payload: str) -> None:
        if not payload:
            return
        for part in re.split(r";|,", payload):
            part = part.strip()
            if not part:
                continue
            if ":" in part:
                key, value = part.split(":", 1)
                rule[key.strip()] = self._parse_value(value.strip())
            elif part.lower().startswith("format "):
                rule["format"] = part[7:].strip()
            else:
                rule[part] = True

    @staticmethod
    def _parse_value(text: str):
        if not text:
            return ""
        text = text.strip()
        if (text.startswith('"') and text.endswith('"')) or (text.startswith("'") and text.endswith("'")):
            return text[1:-1]
        if text.isdigit():
            return int(text)
        try:
            return float(text)
        except ValueError:
            return text
