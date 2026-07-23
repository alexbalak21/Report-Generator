import ast
import operator
import re
from itertools import zip_longest

import openpyxl


class ExcelReader:
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None
        self.formula_workbook = None
        # Cache: { sheet_name: [row_dict, ...] }
        self._sheets: dict[str, list[dict]] = {}
        # Explicit data sheet name (set via set_data_sheet); falls back to active
        self._data_sheet_name: str | None = None

    def set_data_sheet(self, sheet_name: str | None) -> None:
        """Pin the primary data sheet used by get_columns / get_row_as_dict."""
        self._data_sheet_name = sheet_name.strip() if sheet_name else None

    def load(self):
        """Load the workbook and pre-cache all sheets."""
        self._load_workbooks()
        self._cache_sheets()

    def _load_workbooks(self):
        self.workbook = openpyxl.load_workbook(self.filepath, data_only=True)
        self.formula_workbook = openpyxl.load_workbook(self.filepath, data_only=False)

    def _cache_sheets(self):
        self._sheets = {}
        for name in self.workbook.sheetnames:
            self._sheets[name] = self._read_sheet(
                self.workbook[name], self.formula_workbook[name]
            )

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _read_sheet(self, data_sheet, formula_sheet) -> list[dict]:
        """Return all data rows of a sheet as a list of dicts keyed by header."""
        rows = list(data_sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [self._normalize_header(value) for value in rows[0]]
        return [
            self._build_row_dict(headers, data_row, formula_row, data_sheet, formula_sheet)
            for data_row, formula_row in self._iter_sheet_rows(data_sheet, formula_sheet, rows)
        ]

    def _iter_sheet_rows(self, data_sheet, formula_sheet, rows):
        formula_rows = list(formula_sheet.iter_rows())
        yield from zip_longest(rows[1:], formula_rows[1:], fillvalue=())

    def _build_row_dict(self, headers, data_row, formula_row, data_sheet, formula_sheet):
        row_data = {}
        for index, header in enumerate(headers):
            data_value = data_row[index] if index < len(data_row) else None
            formula_cell = formula_row[index] if index < len(formula_row) else None
            value = self._resolve_cell_value(
                data_value, formula_cell, data_sheet, formula_sheet
            )
            row_data[header] = "" if value is None else value
        return row_data

    def _resolve_cell_value(self, data_value, formula_cell, data_sheet, formula_sheet):
        if formula_cell is None or formula_cell.data_type != "f":
            return data_value
        if data_value is not None:
            return data_value
        formula = formula_cell.value
        if not isinstance(formula, str) or not formula.startswith("="):
            return data_value
        try:
            return self._evaluate_formula(formula, data_sheet, formula_sheet, {})
        except Exception:
            return data_value

    def _evaluate_formula(self, formula, data_sheet, formula_sheet, cache):
        expression = self._extract_formula_expression(formula)
        if not expression:
            raise ValueError("Empty formula")

        evaluated = re.sub(
            r"\$?[A-Za-z]{1,3}\$?\d+",
            lambda match: self._resolve_formula_reference(match.group(0), data_sheet, formula_sheet, cache),
            expression,
        )
        return self._safe_eval(evaluated)

    @staticmethod
    def _extract_formula_expression(formula):
        return formula.lstrip("=").strip()

    def _resolve_formula_reference(self, coord, data_sheet, formula_sheet, cache):
        coord = coord.replace("$", "")
        if coord in cache:
            value = cache[coord]
        else:
            value = self._get_cell_value(coord, data_sheet, formula_sheet, cache)
            cache[coord] = value
        return self._format_value_for_expression(value)

    def _get_cell_value(self, coord, data_sheet, formula_sheet, cache):
        cell = data_sheet[coord]
        value = cell.value
        if value is None:
            formula_cell = formula_sheet[coord]
            if formula_cell.data_type == "f" and isinstance(formula_cell.value, str):
                value = self._evaluate_formula(formula_cell.value, data_sheet, formula_sheet, cache)
        return value

    @staticmethod
    def _format_value_for_expression(value):
        if value is None:
            return "0"
        if isinstance(value, bool):
            return str(int(value))
        if isinstance(value, (int, float)):
            return str(value)
        text = str(value).strip().replace(",", ".")
        if text == "":
            return "0"
        try:
            float(text)
            return text
        except ValueError:
            raise ValueError(f"Cannot evaluate formula reference value: {value}")

    @staticmethod
    def _safe_eval(expression):
        operators = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.USub: operator.neg,
            ast.UAdd: operator.pos,
            ast.Pow: operator.pow,
            ast.Mod: operator.mod,
        }

        def _eval(node):
            if isinstance(node, ast.Expression):
                return _eval(node.body)
            if isinstance(node, ast.Constant):
                if isinstance(node.value, (int, float)):
                    return node.value
                raise ValueError("Unsupported constant type")
            if isinstance(node, ast.UnaryOp) and type(node.op) in operators:
                return operators[type(node.op)](_eval(node.operand))
            if isinstance(node, ast.BinOp) and type(node.op) in operators:
                left = _eval(node.left)
                right = _eval(node.right)
                return operators[type(node.op)](left, right)
            raise ValueError(f"Unsupported expression: {ast.dump(node)}")

        parsed = ast.parse(expression, mode="eval")
        return _eval(parsed)

    def _ensure_loaded(self):
        if self.workbook is None:
            self.load()

    # ------------------------------------------------------------------
    # Public API — single active sheet (backwards compatible)
    # ------------------------------------------------------------------

    def _active_sheet_name(self) -> str:
        """Return the pinned data sheet name, or fall back to workbook's active sheet."""
        if self._data_sheet_name:
            # Prefer exact match, then case-insensitive
            if self._data_sheet_name in self._sheets:
                return self._data_sheet_name
            for name in self._sheets:
                if name.lower() == self._data_sheet_name.lower():
                    return name
        return self.workbook.active.title

    @property
    def sheet(self):
        """Return the data sheet (pinned or active), for backwards compatibility."""
        self._ensure_loaded()
        return self.workbook[self._active_sheet_name()]

    @property
    def worksheet(self):
        """Alias — used by processors that need direct cell access."""
        return self.sheet

    def get_columns(self) -> list[str]:
        """Return normalized column names from the data sheet."""
        self._ensure_loaded()
        rows = self._sheets.get(self._active_sheet_name(), [])
        return list(rows[0].keys()) if rows else []

    def get_row_as_dict(self, row_number: int) -> dict:
        """
        Return a row from the data sheet as a dict.
        row_number is 1-based Excel row (row 1 = header, data starts at 2).
        """
        self._ensure_loaded()
        sheet_name = self._active_sheet_name()
        rows = self._sheets.get(sheet_name, [])
        data_index = row_number - 2  # row 2 → index 0
        if data_index < 0 or data_index >= len(rows):
            raise IndexError(f"Row {row_number} out of range in sheet '{sheet_name}'.")
        return rows[data_index]

    # ------------------------------------------------------------------
    # Public API — multi-sheet
    # ------------------------------------------------------------------

    def get_sheet(self, sheet_name: str) -> list[dict]:
        """
        Return all data rows of a named sheet as a list of dicts.
        Raises KeyError if the sheet does not exist.
        """
        self._ensure_loaded()
        if sheet_name not in self._sheets:
            available = ", ".join(self._sheets.keys())
            raise KeyError(
                f"Sheet '{sheet_name}' not found. Available sheets: {available}"
            )
        return self._sheets[sheet_name]

    def find_row(self, sheet_name: str, key_column: str, key_value) -> dict | None:
        """
        Return the first row in sheet_name where key_column == key_value.
        Returns None if no match is found.
        Comparison is done as strings (stripped) to handle mixed int/str Excel values.
        """
        rows = self.get_sheet(sheet_name)
        target = str(key_value).strip().lower() if key_value is not None else ""
        for row in rows:
            cell = row.get(key_column)
            cell_str = str(cell).strip().lower() if cell is not None else ""
            if cell_str == target:
                return row
        return None