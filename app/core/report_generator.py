import os
import re
try:
    import openpyxl
except Exception:
    openpyxl = None

from .excel_reader import ExcelReader
from .word_processor import WordProcessor
from .mapping_loader import MappingLoader
from .state_manager import ReportStateManager
from . import processors

# Characters illegal in Windows filenames
_ILLEGAL_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# Fallback date format when none is specified in mapping config
DEFAULT_DATE_FORMAT = "%d/%m/%Y"

# Column written back to Excel after generation
RAPPORT_GENERE_COLUMN  = "nom rapport generé"
NUMERO_RAPPORT_COLUMN  = "numero rapport"


class ReportGenerator:
    def __init__(self, excel_path, template_path, mapping_path):
        self.excel_path    = excel_path
        self.template_path = template_path
        self.mapping_path  = mapping_path

        self.mapping_loader = MappingLoader(mapping_path)
        self.state_manager  = ReportStateManager(
            os.path.join(os.path.dirname(mapping_path), "report_state.json")
        )
        if openpyxl is None:
            raise ImportError("openpyxl is required for ReportGenerator")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _normalize_header(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        sanitized = _ILLEGAL_FILENAME_CHARS.sub("_", name)
        sanitized = sanitized.strip(". ")
        return sanitized or "report"

    @staticmethod
    def _normalize_field_value(value, date_format: str = DEFAULT_DATE_FORMAT) -> str:
        import datetime as dt
        if isinstance(value, (dt.datetime, dt.date)):
            return value.strftime(date_format)
        if value is None:
            return ""
        return str(value).strip()

    def _build_operations(self, excel_reader: ExcelReader, report_prefix: str = "", row_number: int = 0) -> dict:
        """Build the operations map, injecting excel_reader and report_prefix."""
        return {
            "today":             processors.op_today,
            "uppercase":         processors.op_uppercase,
            "lowercase":         processors.op_lowercase,
            "format":            processors.op_format,
            "concat":            processors.op_concat,
            "report_number":     lambda rule, row: self.state_manager.generate_report_number(),
            "report_prefix":     lambda rule, row: report_prefix,
            "excel_day_counter": lambda rule, row: processors.op_excel_day_counter(
                                     rule, row, excel_reader, row_number),
            "lookup":            lambda rule, row: processors.op_lookup(rule, row, excel_reader),
            "lookup_join":       lambda rule, row: processors.op_lookup_join(rule, row, excel_reader),
        }

    def compute_value(self, rule: dict, row_data: dict, operations: dict):
        operation = rule.get("operation")
        func = operations.get(operation)
        return func(rule, row_data) if func else None

    def _get_placeholder_name(self, placeholder):
        if not isinstance(placeholder, str):
            return None
        match = re.search(r"\{\{(.*?)\}\}", placeholder)
        return match.group(1).strip() if match else None

    def _load_excel(self, row_number: int, data_sheet: str | None = None):
        excel = ExcelReader(self.excel_path)
        if data_sheet:
            excel.set_data_sheet(data_sheet)
        excel.load()
        raw_row = excel.get_row_as_dict(row_number)
        excel_columns = {self._normalize_header(c) for c in excel.get_columns()}
        return excel, raw_row, excel_columns

    def _load_mapping(self):
        config = self.mapping_loader.load_config()
        mapping = self.mapping_loader.load()
        return config, mapping

    def _normalize_row_data(self, raw_row: dict, date_format: str) -> dict:
        return {
            k: self._normalize_field_value(v, date_format)
            for k, v in raw_row.items()
        }

    def _populate_legacy_mappings(
        self,
        mapping: dict,
        excel_columns: set,
        row_data: dict,
        raw_row: dict,
        available_placeholders: set,
        computed_values: dict,
    ) -> dict:
        filled = {}
        for key, rule in mapping.items():
            if not isinstance(rule, str):
                continue

            normalized_key = self._normalize_header(key)
            if normalized_key not in excel_columns:
                continue
            if normalized_key not in row_data:
                continue

            match = re.search(r"\{\{(.*?)\}\}", rule)
            if not match:
                continue

            placeholder_name = match.group(1).strip()
            if placeholder_name not in available_placeholders:
                continue

            value = row_data[normalized_key]
            filled[rule] = value
            computed_values[key] = value
        return filled

    def _populate_column_mappings(
        self,
        mapping: dict,
        excel_columns: set,
        row_data: dict,
        raw_row: dict,
        available_placeholders: set,
        computed_values: dict,
        excel_reader: ExcelReader,
    ) -> dict:
        filled = {}
        for key, rule in mapping.items():
            if not isinstance(rule, dict) or "column" not in rule or "placeholder" not in rule:
                continue

            column_name = self._normalize_header(rule.get("column", ""))
            if column_name not in excel_columns:
                continue

            if rule.get("operations"):
                value = raw_row.get(column_name, row_data.get(column_name, ""))
            else:
                value = row_data.get(column_name, raw_row.get(column_name, ""))

            try:
                value_processed = processors.apply_operations(
                    value, rule.get("operations"), {**row_data, **computed_values}, excel_reader
                )
            except Exception:
                value_processed = value

            placeholder = rule.get("placeholder")
            placeholder_name = self._get_placeholder_name(placeholder)
            if placeholder_name is None:
                continue
            if placeholder_name not in available_placeholders:
                continue

            filled[placeholder] = value_processed
            computed_values[key] = value_processed
        return filled

    def _resolve_computed_fields(
        self,
        mapping: dict,
        row_data: dict,
        operations: dict,
        available_placeholders: set,
        computed_values: dict,
    ) -> dict:
        filled = {}
        max_passes = len(mapping) + 1
        for _ in range(max_passes):
            resolved_any = False
            for key, rule in mapping.items():
                if not isinstance(rule, dict):
                    continue
                if key == "file_name":
                    continue
                if key in computed_values:
                    continue

                enriched = {**row_data, **computed_values}
                try:
                    value = self.compute_value(rule, enriched, operations)
                except Exception:
                    value = None

                if value is None:
                    continue

                computed_values[key] = str(value)
                resolved_any = True

                placeholder_full = f"{{{{{key}}}}}"
                if key in available_placeholders:
                    filled[placeholder_full] = str(value)

            if not resolved_any:
                break
        return filled

    def _resolve_file_name(self, mapping: dict, row_data: dict, operations: dict, computed_values: dict) -> str:
        file_name_rule = mapping.get("file_name")
        if not isinstance(file_name_rule, dict):
            return ""

        enriched = {**row_data, **computed_values}
        try:
            raw_name = self.compute_value(file_name_rule, enriched, operations)
            if raw_name:
                return self._sanitize_filename(str(raw_name))
        except Exception:
            pass
        return ""

    def _populate_report_number_placeholders(self, filled: dict, computed_values: dict) -> None:
        num_val = computed_values.get("sample_number", computed_values.get("numero_rapport", ""))
        if num_val:
            filled["{{numéro_rapport}}"] = num_val
            filled["{{sample_number}}"] = num_val

    def _write_output_document(self, filled: dict, output_path: str) -> None:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        word = WordProcessor(self.template_path)
        word.fill_placeholders(filled, output_path)
    # ------------------------------------------------------------------
    # Write-back: stamp "rapport generé" column in Excel
    # ------------------------------------------------------------------

    def _write_rapport_genere(self, row_number: int, numero_rapport: str, full_numero: str) -> None:
        """
        Write back to Excel:
          - 'numero rapport'     <- yymmdd-N  (e.g. 260621-8)
          - 'nom rapport genere' <- full report name (e.g. NOVOCIB 260621-8)
        Creates columns if absent. row_number is 1-based Excel row.
        """
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

            headers = [
                self._normalize_header(ws.cell(row=1, column=c).value)
                for c in range(1, ws.max_column + 1)
            ]

            def _get_or_create_col(name):
                if name in headers:
                    return headers.index(name) + 1
                col = ws.max_column + 1
                ws.cell(row=1, column=col, value=name)
                headers.append(name)
                return col

            num_col  = _get_or_create_col(NUMERO_RAPPORT_COLUMN)
            full_col = _get_or_create_col(RAPPORT_GENERE_COLUMN)

            ws.cell(row=row_number, column=num_col,  value=numero_rapport)
            ws.cell(row=row_number, column=full_col, value=full_numero)
            wb.save(self.excel_path)
        except Exception as exc:
            print(f"[warn] Could not write back to Excel: {exc}")


    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def generate(self, row_number: int, output_path: str) -> str:
        config, mapping = self._load_mapping()
        excel, raw_row, excel_columns = self._load_excel(
            row_number, data_sheet=config.get("data_sheet")
        )

        date_format = config.get("date_format", DEFAULT_DATE_FORMAT)
        report_prefix = config.get("report_prefix", "")

        row_data = self._normalize_row_data(raw_row, date_format)
        operations = self._build_operations(excel, report_prefix, row_number)

        word = WordProcessor(self.template_path)
        available_placeholders = set(word.extract_placeholders())

        computed_values: dict[str, str] = {"report_prefix": report_prefix}
        filled: dict[str, str] = {}

        filled.update(self._populate_legacy_mappings(
            mapping,
            excel_columns,
            row_data,
            raw_row,
            available_placeholders,
            computed_values,
        ))

        filled.update(self._populate_column_mappings(
            mapping,
            excel_columns,
            row_data,
            raw_row,
            available_placeholders,
            computed_values,
            excel,
        ))

        filled.update(self._resolve_computed_fields(
            mapping,
            row_data,
            operations,
            available_placeholders,
            computed_values,
        ))

        full_name = self._resolve_file_name(mapping, row_data, operations, computed_values)

        self._populate_report_number_placeholders(filled, computed_values)
        self._write_output_document(filled, output_path)

        # ── Write back to Excel ───────────────────────────────────────
        # numero_rapport / sample_number = yymmdd-N (e.g. 260621-8)
        # file_name resolved = full name for "nom rapport generé" column
        numero = computed_values.get("sample_number", computed_values.get("numero_rapport", ""))
        if numero:
            self._write_rapport_genere(row_number, numero, full_name or numero)

        return output_path